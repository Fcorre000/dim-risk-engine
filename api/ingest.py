import pandas as pd
import numpy as np
import io
import zipfile
import openpyxl
from typing import Union

# XLSX safety ceilings — each defends a different zip-bomb vector.
# Reasonable real FedEx exports stay well under all of these; tune down if needed.
MAX_XLSX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024   # 250 MB total uncompressed
MAX_XLSX_ENTRY_BYTES        = 150 * 1024 * 1024   # 150 MB single entry
MAX_XLSX_ENTRIES            = 200                 # archive part count
MAX_XLSX_COMPRESS_RATIO     = 100                 # per-entry bytes-out / bytes-in


def _assert_xlsx_safe(data: bytes) -> None:
    """Raise ValueError if the XLSX archive looks like a zip bomb.

    Checks we apply BEFORE handing the bytes to openpyxl:
    - archive opens cleanly as a zip
    - entry count is bounded (prevents part-flood DoS)
    - no single entry decompresses larger than `MAX_XLSX_ENTRY_BYTES`
    - per-entry compression ratio stays under `MAX_XLSX_COMPRESS_RATIO`
    - cumulative uncompressed size stays under `MAX_XLSX_UNCOMPRESSED_BYTES`

    All five come from the ZIP header without actually decompressing, so this
    is O(N_entries) and safe to run on untrusted input. The raw file-size cap
    in the API layer bounds the compressed input; these bounds cover expansion.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            infos = zf.infolist()
            if len(infos) > MAX_XLSX_ENTRIES:
                raise ValueError("XLSX has too many internal entries")
            total_uncompressed = 0
            for info in infos:
                if info.file_size > MAX_XLSX_ENTRY_BYTES:
                    raise ValueError("XLSX entry too large (possible zip bomb)")
                if info.compress_size > 0:
                    ratio = info.file_size / info.compress_size
                    if ratio > MAX_XLSX_COMPRESS_RATIO:
                        raise ValueError("XLSX compression ratio too high (possible zip bomb)")
                total_uncompressed += info.file_size
                if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError("XLSX total uncompressed size exceeds limit")
    except zipfile.BadZipFile as exc:
        raise ValueError("Invalid XLSX file") from exc

FEATURE_COLS = [
    'Original Weight (Pounds)', 'Dimmed Height (cm)', 'Dimmed Width (cm)',
    'Dimmed Length (cm)', 'volume', 'dim_weight_calculator', 'dim_weight_ratio',
    'has_dimensions', 'billable_weight', 'billable_weight_ceil',
    'ship_year', 'ship_month', 'months_since_start',
    'Service Type_CTAG', 'Service Type_ES', 'Service Type_FO',
    'Service Type_MWT', 'Service Type_ON', 'Service Type_PO', 'Service Type_QH',
    'Service Type_RMGR', 'Service Type_RW', 'Service Type_S7', 'Service Type_S8',
    'Service Type_SG', 'Service Type_SO', 'Service Type_TA', 'Service Type_XS',
    'Pay Type_Bill_Recipient', 'Pay Type_Bill_Sender_Prepaid', 'Pay Type_Bill_Third_Party',
    'Pay Type_Other4',
    'zone_clean_02', 'zone_clean_03', 'zone_clean_04', 'zone_clean_05',
    'zone_clean_06', 'zone_clean_07', 'zone_clean_08', 'zone_clean_09',
    'zone_clean_17', 'zone_clean_Other'
]

LEAKAGE_COLS = ["Shipment Rated Weight (Pounds)", "Net Charge Billed Currency"]

REQUIRED_COLS = [
    "Tracking Number", "Original Weight (Pounds)",
    "Dimmed Height (cm)", "Dimmed Width (cm)", "Dimmed Length (cm)",
    "Service Type", "Pay Type",
    "Pricing Zone", "Shipment DIM Flag (Y or N)", "Net Charge Billed Currency",
]


def parse_invoice(file_bytes: io.BytesIO, filename: str) -> pd.DataFrame:
    """Parse an uploaded invoice file (.xlsx or .csv) and return a validated DataFrame.

    Args:
        file_bytes: File content as BytesIO.
        filename: Original filename — used to detect .xlsx vs .csv.

    Returns:
        DataFrame with raw invoice rows (leakage columns still present).

    Raises:
        ValueError: If file type unsupported or required columns are missing.
    """
    fn = filename.lower()
    if fn.endswith(".xlsx"):
        # Buffer the bytes once so we can zip-bomb-check before openpyxl touches them.
        data = file_bytes.read()
        _assert_xlsx_safe(data)
        df = pd.read_excel(io.BytesIO(data), engine="openpyxl")
    elif fn.endswith(".csv"):
        df = pd.read_csv(file_bytes)
    else:
        raise ValueError("Unsupported file type. Upload .xlsx or .csv")

    # Normalize tracking number column — FedEx exports use different names
    for alt in ("Shipment Tracking Number", "Master Tracking Number"):
        if alt in df.columns and "Tracking Number" not in df.columns:
            df = df.rename(columns={alt: "Tracking Number"})
            break

    # Customs Value is optional — fill if missing
    if "Customs Value" not in df.columns:
        df["Customs Value"] = 0.0

    # Validate required columns (excluding Customs Value which we just filled)
    required_to_check = [c for c in REQUIRED_COLS if c != "Customs Value"]
    missing = [c for c in required_to_check if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    # Drop NonTrans rows — matches training behavior (non-transportation charges)
    if "Service Type" in df.columns:
        df = df[df["Service Type"] != "NonTrans"].reset_index(drop=True)

    return df


def parse_invoice_chunks(file_obj, filename: str, chunksize: int = 1000):
    """Generator that yields preprocessed DataFrame chunks for CSV and XLSX files.

    Args:
        file_obj: File-like object (e.g., SpooledTemporaryFile from FastAPI).
        filename: Original filename — used to detect .xlsx vs .csv.
        chunksize: Number of rows per chunk.

    Yields:
        DataFrame chunks with the same columns as parse_invoice would return.

    Raises:
        ValueError: If file type unsupported or required columns are missing.
    """
    fn = filename.lower()

    def _preprocess_chunk(chunk: pd.DataFrame) -> pd.DataFrame:
        # Normalize tracking number column
        for alt in ("Shipment Tracking Number", "Master Tracking Number"):
            if alt in chunk.columns and "Tracking Number" not in chunk.columns:
                chunk = chunk.rename(columns={alt: "Tracking Number"})
                break

        # Customs Value is optional — fill if missing
        if "Customs Value" not in chunk.columns:
            chunk["Customs Value"] = 0.0

        # Validate required columns (excluding Customs Value which we just filled)
        required_to_check = [c for c in REQUIRED_COLS if c != "Customs Value"]
        missing = [c for c in required_to_check if c not in chunk.columns]
        if missing:
            raise ValueError(f"Missing required columns: {missing}")

        # Drop NonTrans rows
        if "Service Type" in chunk.columns:
            chunk = chunk[chunk["Service Type"] != "NonTrans"].reset_index(drop=True)

        return chunk

    if fn.endswith(".csv"):
        for chunk in pd.read_csv(file_obj, chunksize=chunksize):
            yield _preprocess_chunk(chunk)

    elif fn.endswith(".xlsx"):
        # Buffer once so we can zip-bomb-check before handing to openpyxl.
        # file_obj may be a SpooledTemporaryFile or BytesIO — both support .read().
        xlsx_bytes = file_obj.read()
        _assert_xlsx_safe(xlsx_bytes)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(min_row=1, max_row=1)
            header = [cell.value for cell in next(rows_iter)]

            # Apply tracking number alias normalization to header
            for alt in ("Shipment Tracking Number", "Master Tracking Number"):
                if alt in header and "Tracking Number" not in header:
                    header[header.index(alt)] = "Tracking Number"
                    break

            batch = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                batch.append(row)
                if len(batch) >= chunksize:
                    chunk = pd.DataFrame(batch, columns=header)
                    yield _preprocess_chunk(chunk)
                    batch = []

            # Yield remaining rows
            if batch:
                chunk = pd.DataFrame(batch, columns=header)
                yield _preprocess_chunk(chunk)
        finally:
            wb.close()

    else:
        raise ValueError("Unsupported file type. Upload .xlsx or .csv")


def clean_zone(raw_zone) -> str:
    """Normalize a FedEx Pricing Zone value to a zero-padded string or 'Other'.

    Valid zones are integers 1–50 (inclusive). Anything outside that range,
    non-numeric, or None maps to 'Other'.

    Examples:
        clean_zone("2")   -> "02"
        clean_zone(17)    -> "17"
        clean_zone("A")   -> "Other"
        clean_zone("99")  -> "Other"
        clean_zone(None)  -> "Other"
        clean_zone(" 8 ") -> "08"
    """
    try:
        z = int(float(str(raw_zone).strip()))
        if 1 <= z <= 50:
            return f"{z:02d}"
        else:
            return "Other"
    except (ValueError, TypeError):
        return "Other"


def build_feature_matrix(df: pd.DataFrame) -> pd.DataFrame:
    """Transform a raw invoice DataFrame into the exact 42-column feature matrix
    expected by the XGBoost models.

    Steps:
    1. Engineer volume, dim_weight_calculator, dim_weight_ratio, has_dimensions
    2. Normalize Pricing Zone via clean_zone, then one-hot encode
    3. One-hot encode Service Type and Pay Type
    4. Reindex to FEATURE_COLS (fills missing columns with 0, drops extras)

    Leakage columns (LEAKAGE_COLS) are never included in the output.

    Args:
        df: Raw invoice DataFrame from parse_invoice.

    Returns:
        DataFrame with exactly 42 columns matching FEATURE_COLS.
    """
    df = df.copy()

    # --- Engineered features ---
    # Convert cm to inches to match training preprocessing (FedEx DIM divisor 139 = in³/lb)
    CM_PER_IN = 2.54
    H = df["Dimmed Height (cm)"].fillna(0) / CM_PER_IN
    W = df["Dimmed Width (cm)"].fillna(0) / CM_PER_IN
    L = df["Dimmed Length (cm)"].fillna(0) / CM_PER_IN

    df["volume"] = H * W * L
    df["dim_weight_calculator"] = df["volume"] / 139.0

    orig_weight = df["Original Weight (Pounds)"].replace(0, np.nan)
    df["dim_weight_ratio"] = df["dim_weight_calculator"] / orig_weight
    df["dim_weight_ratio"] = df["dim_weight_ratio"].replace([np.inf, -np.inf], 0).fillna(0)

    df["has_dimensions"] = ((H > 0) & (W > 0) & (L > 0)).astype(int)

    # Billable weight: max of original weight and DIM weight
    df["billable_weight"] = np.maximum(
        df["Original Weight (Pounds)"].fillna(0),
        df["dim_weight_calculator"],
    )
    df["billable_weight_ceil"] = np.ceil(df["billable_weight"])

    # --- Temporal features from Shipment Date ---
    ship_date = None
    for col_name in ("Shipment Date (mm/dd/yyyy)", "Shipment Date"):
        if col_name in df.columns:
            ship_date = pd.to_datetime(df[col_name], errors="coerce")
            break
    if ship_date is not None:
        df["ship_year"] = ship_date.dt.year
        df["ship_month"] = ship_date.dt.month
        # months_since_start: months elapsed since April 2024 (training data start)
        df["months_since_start"] = (ship_date.dt.year - 2024) * 12 + ship_date.dt.month
    else:
        df["ship_year"] = 0
        df["ship_month"] = 0
        df["months_since_start"] = 0

    # --- Zone normalization and OHE ---
    df["zone_clean"] = df["Pricing Zone"].apply(clean_zone)
    zone_dummies = pd.get_dummies(df["zone_clean"], prefix="zone_clean")

    # --- Service Type and Pay Type OHE ---
    svc_dummies = pd.get_dummies(df["Service Type"], prefix="Service Type")
    pay_dummies = pd.get_dummies(df["Pay Type"], prefix="Pay Type")

    # --- Assemble feature DataFrame ---
    feature_df = pd.concat([df, zone_dummies, svc_dummies, pay_dummies], axis=1)

    # Reindex to exact FEATURE_COLS — fills missing columns with 0, drops leakage and extras
    return feature_df.reindex(columns=FEATURE_COLS, fill_value=0)
