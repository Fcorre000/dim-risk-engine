from collections import defaultdict
from contextlib import asynccontextmanager
import io
import logging
import os
import pathlib
import time
import warnings
from typing import Optional

import xgboost as xgb
from fastapi import FastAPI, File, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel

import json as _json

from ingest import parse_invoice, parse_invoice_chunks
from inference import run_inference, load_residual_quantiles

warnings.filterwarnings("ignore", message=".*Booster.*", category=UserWarning)

logger = logging.getLogger("dimrisk.api")

# Generic client-facing message for unexpected parse failures. Real details
# stay in server logs so debugging doesn't require handing attackers a map
# of our pandas/openpyxl stack (column guesses, encoding fallbacks, etc.).
_GENERIC_PARSE_ERROR = "Could not parse invoice file. Ensure it is a valid FedEx .xlsx or .csv export."

MAX_FILE_BYTES = 50 * 1024 * 1024  # 50 MB
_READ_CHUNK = 64 * 1024            # streaming read granularity

RATE_LIMIT = 10        # max requests per IP
RATE_WINDOW = 60.0     # per 60 seconds
_RATE_MAX_KEYS = 10_000  # hard cap on distinct IPs tracked — prevents memory leak
_SWEEP_INTERVAL = RATE_WINDOW  # evict stale IPs at most once per window

_rate_limit_store: dict[str, list[float]] = defaultdict(list)
_last_sweep: float = 0.0


def _client_ip(request: Request) -> str:
    """Resolve the real client IP behind a reverse proxy (Render, CF, etc.).

    `request.client.host` is the proxy's IP, not the caller's, so rate limits
    keyed off it collapse every user into a single bucket. We prefer the
    leftmost entry of `X-Forwarded-For` (the real client as seen by the edge),
    fall back to `X-Real-IP`, and finally to the raw socket peer.

    NOTE: `X-Forwarded-For` is *client-controllable* — a malicious caller can
    prepend their own entry to forge a different key. Render's LB appends the
    real IP, so the leftmost-entry strategy trusts the edge to overwrite. If
    this app ever moves off a single-hop proxy, revisit this with a
    configurable trusted-proxy depth.
    """
    xff = request.headers.get("x-forwarded-for")
    if xff:
        first = xff.split(",", 1)[0].strip()
        if first:
            return first
    real_ip = request.headers.get("x-real-ip")
    if real_ip:
        return real_ip.strip()
    return request.client.host if request.client else "unknown"


def _sweep_rate_store(now: float) -> None:
    """Drop IPs whose last activity fell outside the current rate window.

    Called opportunistically on each rate-limit check but throttled to at most
    one full sweep per window so the amortised cost stays O(1) per request.
    """
    global _last_sweep
    if now - _last_sweep < _SWEEP_INTERVAL:
        return
    _last_sweep = now
    cutoff = now - RATE_WINDOW
    stale = [k for k, ts in _rate_limit_store.items() if not ts or ts[-1] < cutoff]
    for k in stale:
        del _rate_limit_store[k]


def check_rate_limit(ip: str) -> bool:
    now = time.time()
    _sweep_rate_store(now)

    # Safety net: if the dict has ballooned past the cap (e.g. an attacker
    # cycling through spoofed X-Forwarded-For values faster than the sweep),
    # evict the oldest-touched entries so memory doesn't grow unbounded.
    if len(_rate_limit_store) > _RATE_MAX_KEYS:
        overflow = len(_rate_limit_store) - _RATE_MAX_KEYS
        # Sort by most-recent timestamp ascending, drop the oldest `overflow`.
        oldest = sorted(
            _rate_limit_store.items(),
            key=lambda kv: kv[1][-1] if kv[1] else 0.0,
        )[:overflow]
        for k, _ in oldest:
            _rate_limit_store.pop(k, None)

    window_start = now - RATE_WINDOW
    timestamps = [t for t in _rate_limit_store.get(ip, ()) if t > window_start]
    if len(timestamps) >= RATE_LIMIT:
        _rate_limit_store[ip] = timestamps
        return False
    timestamps.append(now)
    _rate_limit_store[ip] = timestamps
    return True


async def _read_bounded(file: UploadFile, limit: int) -> Optional[bytes]:
    """Read the upload up to `limit` bytes, aborting early on overflow.

    Returns the bytes on success, or None if the stream exceeded `limit`
    (in which case the caller should emit HTTP 413 without materialising
    the rest of the body). This matters: the previous code did an unbounded
    `await file.read()` and only checked size *after* the whole payload
    was already resident, making the 50 MB cap a client-visible UX detail
    rather than a server-side resource guard.
    """
    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = await file.read(_READ_CHUNK)
        if not chunk:
            break
        total += len(chunk)
        if total > limit:
            return None
        chunks.append(chunk)
    return b"".join(chunks)


@asynccontextmanager
async def lifespan(app: FastAPI):
    model_path_env = os.environ.get("MODEL_PATH")
    if model_path_env:
        models_dir = pathlib.Path(model_path_env)
    else:
        models_dir = pathlib.Path(__file__).parent.parent / "models"

    # Native XGBoost UBJ format — plain data, no code execution on load.
    # Migrated from pickle.load in 2026-04-18 to close an RCE-on-boot sink.
    # Regenerate with: scripts/convert_models_to_ubj.py
    clf = xgb.Booster()
    clf.load_model(str(models_dir / "xgb_classifier.ubj"))
    reg = xgb.Booster()
    reg.load_model(str(models_dir / "xgb_regressor.ubj"))
    app.state.clf = clf
    app.state.reg = reg
    app.state.residual_quantiles = load_residual_quantiles(models_dir)
    yield


app = FastAPI(title="DimRisk Engine", lifespan=lifespan)

# CORS: fail-closed. CORS_ORIGINS must be a comma-separated list of exact origins.
# If unset, default to local Vite dev servers only — never wildcard, and never
# combine with allow_credentials (the API has no cookies/sessions anyway).
_raw_origins = os.environ.get("CORS_ORIGINS", "").strip()
if _raw_origins:
    _allow_origins = [o.strip() for o in _raw_origins.split(",") if o.strip()]
else:
    _allow_origins = ["http://localhost:5173", "http://localhost:4173"]
if "*" in _allow_origins:
    raise RuntimeError(
        "CORS_ORIGINS must not contain '*'. List exact origins (e.g. "
        "https://app.example.com,https://staging.example.com)."
    )

app.add_middleware(
    CORSMiddleware,
    allow_origins=_allow_origins,
    allow_credentials=False,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["Content-Type"],
)


# Defence-in-depth security headers. The API returns JSON, not HTML, so XSS
# isn't the primary vector here — these mostly harden error pages, stray
# HTML responses, and anything loaded via <script src> or <iframe src>.
# HSTS is emitted unconditionally: Render terminates TLS so every real
# request reaches the client over HTTPS anyway, and browsers ignore HSTS
# on plain-HTTP responses (so local `http://localhost:8000` dev is fine).
@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
    response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    response.headers.setdefault("Cross-Origin-Opener-Policy", "same-origin")
    # API is consumed cross-origin by the SPA; CORP must allow that.
    response.headers.setdefault("Cross-Origin-Resource-Policy", "cross-origin")
    return response


class ShipmentResult(BaseModel):
    row_index: int                 # globally unique per-upload id (stable React key)
    tracking_number: Optional[str] # nullable — FedEx exports occasionally omit it
    service_type: str              # e.g. "FO", "SG", "PO"
    weight_lbs: float              # Original Weight (Pounds)
    dim_length: float              # Dimmed Length (in)
    dim_width: float               # Dimmed Width (in)
    dim_height: float              # Dimmed Height (in)
    zone: str                      # Pricing Zone, normalized ("02", "Other")
    shipment_date: Optional[str]   # "YYYY-MM-DD" or null if not in source
    dim_flag_probability: float    # P(DIM=Y), 0.0-1.0
    actual_net_charge: float       # dollars, from Net Charge Billed Currency column
    predicted_net_charge: float    # dollars, after np.expm1()
    predicted_net_charge_low: float   # 5th percentile lower bound (dollars)
    predicted_net_charge_high: float  # 95th percentile upper bound (dollars)
    dim_anomaly: Optional[str]     # "Unexpected" or None
    dim_confidence: Optional[float]   # P(DIM=N) when dim_anomaly is Unexpected, else None
    cost_anomaly: Optional[str]    # "Review" or None
    cost_confidence: Optional[str]    # "High" if actual > pred_high, else None


@app.get("/health")
async def health():
    return {
        "status": "ok",
        "models_loaded": hasattr(app.state, "clf") and hasattr(app.state, "reg"),
    }


@app.post("/analyze", response_model=list[ShipmentResult])
async def analyze(request: Request, file: UploadFile = File(...)):
    if not check_rate_limit(_client_ip(request)):
        return JSONResponse(status_code=429, content={"detail": "Too many requests. Try again in a minute."})

    start = time.perf_counter()

    filename = file.filename or "upload.csv"
    # Bounded read: aborts at MAX_FILE_BYTES+1 instead of slurping the whole body first
    contents = await _read_bounded(file, MAX_FILE_BYTES)
    if contents is None:
        return JSONResponse(status_code=413, content={"detail": f"File too large. Maximum size is {MAX_FILE_BYTES // 1024 // 1024} MB."})

    # Parse and preprocess
    try:
        df = parse_invoice(io.BytesIO(contents), filename)
    except ValueError as e:
        # Our own validation errors from ingest.py — safe to surface verbatim
        # (missing-columns / zip-bomb / unsupported-type messages we author).
        logger.info("parse_invoice rejected upload (%s): %s", filename, e)
        return JSONResponse(status_code=422, content={"detail": str(e)})
    except Exception:
        # Anything pandas/openpyxl raises that ISN'T a ValueError we authored
        # gets a generic response — stack details stay in server logs.
        logger.exception("parse_invoice failed unexpectedly (%s)", filename)
        return JSONResponse(status_code=422, content={"detail": _GENERIC_PARSE_ERROR})

    # Run inference
    try:
        results = run_inference(df, request.app.state.clf, request.app.state.reg, request.app.state.residual_quantiles)
    except Exception:
        logger.exception("run_inference failed (%s, rows=%d)", filename, len(df))
        return JSONResponse(status_code=500, content={"detail": "Internal error while scoring invoice."})

    elapsed = time.perf_counter() - start

    return results


@app.post("/analyze/stream")
async def analyze_stream(request: Request, file: UploadFile = File(...)):
    if not check_rate_limit(_client_ip(request)):
        return JSONResponse(status_code=429, content={"detail": "Too many requests. Try again in a minute."})

    filename = file.filename or "upload.csv"
    # Bounded read: short-circuits before committing an oversize body to memory
    contents = await _read_bounded(file, MAX_FILE_BYTES)
    if contents is None:
        return JSONResponse(status_code=413, content={"detail": f"File too large. Maximum size is {MAX_FILE_BYTES // 1024 // 1024} MB."})

    # Cheap row estimate from bytes already in memory
    total = None
    fn = filename.lower()
    try:
        if fn.endswith(".csv"):
            total = max(0, contents.count(b"\n") - 1)  # subtract header row
        elif fn.endswith(".xlsx"):
            import openpyxl
            from ingest import _assert_xlsx_safe
            # Zip-bomb guard — must run before openpyxl parses the archive.
            # Any ValueError here is swallowed by the outer try so row-count
            # stays best-effort; the downstream parse will re-check and surface
            # the error through the streamed __error__ frame.
            _assert_xlsx_safe(contents)
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            ws = wb.active
            total = max(0, ws.max_row - 1) if ws.max_row else None  # subtract header
            wb.close()
    except Exception:
        pass

    clf = request.app.state.clf
    reg = request.app.state.reg
    rq = request.app.state.residual_quantiles

    def generate():
        yield _json.dumps({"__meta__": True, "total": total}) + "\n"
        offset = 0
        try:
            for chunk in parse_invoice_chunks(io.BytesIO(contents), filename, chunksize=1000):
                for row in run_inference(chunk, clf, reg, rq, start_index=offset):
                    yield _json.dumps(row) + "\n"
                offset += len(chunk)
        except ValueError as e:
            logger.info("analyze_stream validation error (%s): %s", filename, e)
            yield _json.dumps({"__error__": str(e)}) + "\n"
        except Exception:
            logger.exception("analyze_stream failed unexpectedly (%s)", filename)
            yield _json.dumps({"__error__": _GENERIC_PARSE_ERROR}) + "\n"

    return StreamingResponse(generate(), media_type="application/x-ndjson")


@app.get("/demo/stream")
async def demo_stream(request: Request):
    """Stream inference results for the built-in 3,000-row sample invoice.

    Response format is identical to /analyze/stream — NDJSON, one object per line,
    first line is the __meta__ row with total count.
    """
    demo_path = pathlib.Path(__file__).parent / "sample_invoice.csv"
    if not demo_path.exists():
        return JSONResponse(status_code=404, content={"detail": "Sample invoice not found on server."})

    contents = demo_path.read_bytes()
    filename = "sample_invoice.csv"

    total = max(0, contents.count(b"\n") - 1)

    clf = request.app.state.clf
    reg = request.app.state.reg
    rq = request.app.state.residual_quantiles

    def generate():
        yield _json.dumps({"__meta__": True, "total": total}) + "\n"
        offset = 0
        try:
            for chunk in parse_invoice_chunks(io.BytesIO(contents), filename, chunksize=1000):
                for row in run_inference(chunk, clf, reg, rq, start_index=offset):
                    yield _json.dumps(row) + "\n"
                offset += len(chunk)
        except ValueError as e:
            logger.info("demo_stream validation error: %s", e)
            yield _json.dumps({"__error__": str(e)}) + "\n"
        except Exception:
            logger.exception("demo_stream failed unexpectedly")
            yield _json.dumps({"__error__": _GENERIC_PARSE_ERROR}) + "\n"

    return StreamingResponse(generate(), media_type="application/x-ndjson")
